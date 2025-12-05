from __future__ import annotations

import csv
import io
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from src.auth.dependencies import get_current_user
from src.db import get_db_session
from src.models import BaselineValue, User
from src.schemas.baseline_value import (
    BaselineValueCreate,
    BaselineValueResponse,
    BaselineValueUpdate,
)

router = APIRouter(prefix="/baseline-values", tags=["baseline_values"])


def _ensure_org_scope(db: Session, current_user: User) -> Any:
    return (
        db.query(BaselineValue)
        .filter(BaselineValue.organization_id == current_user.organization_id)
    )


def _parse_rows(rows: list[dict[str, Any]]) -> list[BaselineValueCreate]:
    parsed: list[BaselineValueCreate] = []
    for index, row in enumerate(rows, start=1):
        try:
            parsed.append(
                BaselineValueCreate(
                    name=str(row.get("name", "")).strip(),
                    metric_key=str(row.get("metric_key", "")).strip(),
                    value=float(row.get("value")),
                    unit=row.get("unit") or None,
                    notes=row.get("notes") or None,
                    building_id=int(row["building_id"]) if row.get("building_id") else None,
                    chiller_unit_id=int(row["chiller_unit_id"]) if row.get("chiller_unit_id") else None,
                )
            )
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid data on row {index}: {exc}",
            ) from exc
    return parsed


@router.get("", response_model=list[BaselineValueResponse])
def list_baseline_values(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db_session),
):
    return _ensure_org_scope(db, current_user).order_by(BaselineValue.id).all()


@router.post("", response_model=BaselineValueResponse, status_code=status.HTTP_201_CREATED)
def create_baseline_value(
    payload: BaselineValueCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db_session),
):
    record = BaselineValue(
        organization_id=current_user.organization_id,
        **payload.model_dump(),
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.patch("/{baseline_id}", response_model=BaselineValueResponse)
def update_baseline_value(
    baseline_id: int,
    payload: BaselineValueUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db_session),
):
    record = _ensure_org_scope(db, current_user).filter(BaselineValue.id == baseline_id).first()
    if record is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Baseline not found")

    update_data = payload.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(record, field, value)
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.delete("/{baseline_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_baseline_value(
    baseline_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db_session),
):
    record = _ensure_org_scope(db, current_user).filter(BaselineValue.id == baseline_id).first()
    if record is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Baseline not found")
    db.delete(record)
    db.commit()
    return None


def _load_csv(file: UploadFile) -> list[dict[str, Any]]:
    content = file.file.read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    return list(reader)


def _load_excel(file: UploadFile) -> list[dict[str, Any]]:
    workbook = load_workbook(file.file, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({header: value for header, value in zip(headers, row)})
    return rows


@router.post("/import")
def import_baseline_values(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db_session),
):
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is required")

    if file.filename.lower().endswith(".csv"):
        rows = _load_csv(file)
    elif file.filename.lower().endswith(".xlsx"):
        rows = _load_excel(file)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file format. Please upload CSV or XLSX.",
        )

    payloads = _parse_rows(rows)
    created_count = 0
    for payload in payloads:
        record = BaselineValue(
            organization_id=current_user.organization_id,
            **payload.model_dump(),
        )
        db.add(record)
        created_count += 1
    db.commit()

    return {"created": created_count}
